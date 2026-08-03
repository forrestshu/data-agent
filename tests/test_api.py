"""HTTP 接口验收：验证 React 所依赖的查询、确认与知识状态契约。"""

from __future__ import annotations

from io import BytesIO
from decimal import Decimal
import time
import unittest
from typing import Any

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from data_agent.api import _json_safe, create_app
from data_agent.data_sources import SQLITE_SOURCE_ID


def create_test_app(**kwargs: Any):
    """API 测试固定使用 SQLite，避免依赖内网 SQL Server。"""

    return create_app(
        source_id=SQLITE_SOURCE_ID,
        persist_source=False,
        **kwargs,
    )


class FakeLLMClient:
    """测试替身：用固定 JSON 和文本验证 AI 编排，不访问真实 DeepSeek 网络。"""

    provider = "deepseek"
    model = "deepseek-v4-flash"

    def __init__(self, intent: dict[str, Any], answer: str | list[str] = "") -> None:
        self.intent = intent
        self.answer = answer
        self.text_calls = 0

    def complete_json(
        self,
        system_prompt: str,
        user_prompt: str,
        max_tokens: int = 1200,
    ) -> dict[str, Any]:
        """返回测试预设意图，并保留与生产客户端一致的方法签名。"""

        return self.intent

    def complete_text(
        self,
        system_prompt: str,
        user_prompt: str,
        max_tokens: int = 1200,
    ) -> str:
        """返回测试预设自然语言回答。"""

        self.text_calls += 1
        if isinstance(self.answer, list):
            index = min(self.text_calls - 1, len(self.answer) - 1)
            return self.answer[index]
        return self.answer


class ColumnLocalizingLLMClient(FakeLLMClient):
    """为未知 SQL 别名返回中文表头，验证动态字段本地化回退。"""

    def __init__(self, intent: dict[str, Any], answer: str) -> None:
        super().__init__(intent, answer)
        self.localization_calls = 0

    def complete_json(
        self,
        system_prompt: str,
        user_prompt: str,
        max_tokens: int = 1200,
    ) -> dict[str, Any]:
        if "尚未登记语义的英文列名" in system_prompt:
            self.localization_calls += 1
            return {"column_labels": {"total_qty": "库存总数量"}}
        return super().complete_json(system_prompt, user_prompt, max_tokens)


class BrokenIntentLLMClient:
    """测试替身：连续返回不完整意图，随后用文本能力生成客户可读失败说明。"""

    provider = "deepseek"
    model = "deepseek-v4-flash"

    def complete_json(
        self,
        system_prompt: str,
        user_prompt: str,
        max_tokens: int = 1200,
    ) -> dict[str, Any]:
        """模拟模型连续漏掉契约字段。"""

        return {"status": "ready"}

    def complete_text(
        self,
        system_prompt: str,
        user_prompt: str,
        max_tokens: int = 1200,
    ) -> str:
        """模拟模型把内部失败整理成客户能理解的自然语言。"""

        return "这次没有完成问题理解，请稍后重新提交；你的数据不会受到影响。"


class TechnicalFailureAnswerLLMClient(BrokenIntentLLMClient):
    """测试替身：模拟模型在错误说明中泄露技术词。"""

    def complete_text(
        self,
        system_prompt: str,
        user_prompt: str,
        max_tokens: int = 1200,
    ) -> str:
        """返回应被后端安全过滤的技术化回答。"""

        return "DeepSeek JSON 接口契约校验失败，请检查 API。"


class RepeatedClarificationLLMClient:
    """测试替身：第一次重复旧问题，收到守卫反馈后改为可执行查询。"""

    provider = "deepseek"
    model = "deepseek-v4-flash"

    def __init__(self) -> None:
        self.json_calls = 0
        self.user_prompts: list[str] = []

    def complete_json(
        self,
        system_prompt: str,
        user_prompt: str,
        max_tokens: int = 1200,
    ) -> dict[str, Any]:
        """先返回重复追问，再模拟模型依据完整历史完成安全修复。"""

        self.json_calls += 1
        self.user_prompts.append(user_prompt)
        if self.json_calls == 1:
            return {
                "status": "clarification_required",
                "confidence": 0.5,
                "intent_summary": "物料查询",
                "route_reason": "需要指标",
                "clarification_question": "你想查库存、采购价格，还是基础资料？",
            }
        return {
            "status": "ready",
            "confidence": 0.95,
            "intent_summary": "查询物料库存",
            "route_reason": "用户已回答要查库存",
            "clarification_question": None,
            "matched_concepts": ["物料", "库存"],
            "source_views": ["AiQueryPartOnHandV"],
            "sql": "SELECT PartNum, Qty FROM AiQueryPartOnHandV WHERE PartNum = ?",
            "parameters": ["110000012"],
            "assumptions": [],
        }

    def complete_text(
        self,
        system_prompt: str,
        user_prompt: str,
        max_tokens: int = 1200,
    ) -> str:
        """返回客户回答，证明重复追问已被内部修复而非继续展示。"""

        return "物料 110000012 当前库存为 8290。"


class InvalidClarificationRepairingLLMClient:
    """测试替身：先提出索要查询答案的无效追问，再按语义守卫反馈改成单视图 SQL。"""

    provider = "deepseek"
    model = "deepseek-v4-flash"

    def __init__(
        self,
        invalid_intent: dict[str, Any],
        repaired_intent: dict[str, Any],
    ) -> None:
        """保存两次模型输出，明确展示“错误追问 -> 自动修复”的编排状态变化。"""

        self.invalid_intent = invalid_intent
        self.repaired_intent = repaired_intent
        self.json_calls = 0
        self.user_prompts: list[str] = []

    def complete_json(
        self,
        system_prompt: str,
        user_prompt: str,
        max_tokens: int = 1200,
    ) -> dict[str, Any]:
        """第一次返回无效澄清，第二次证明后端反馈能够引导模型生成安全查询。"""

        self.json_calls += 1
        self.user_prompts.append(user_prompt)
        return self.invalid_intent if self.json_calls == 1 else self.repaired_intent

    def complete_text(
        self,
        system_prompt: str,
        user_prompt: str,
        max_tokens: int = 1200,
    ) -> str:
        """回答层使用固定文本；本测试只关注理解与澄清边界。"""

        return "已根据明确条件完成查询。"


class ApiTests(unittest.TestCase):
    """接口层测试：通过真实 ASGI 请求覆盖主要前后端数据流。"""

    @classmethod
    def setUpClass(cls) -> None:
        """测试生命周期：启动 FastAPI 并执行其知识画像预检。"""

        cls.client_context = TestClient(
            create_test_app(use_environment_ai=False, require_ai=False)
        )
        cls.client = cls.client_context.__enter__()

    @classmethod
    def tearDownClass(cls) -> None:
        """测试生命周期：关闭 TestClient，释放应用生命周期资源。"""

        cls.client_context.__exit__(None, None, None)

    def test_health_reports_ready_knowledge(self) -> None:
        """健康接口必须同时证明 API 存活且当前知识画像可用。"""

        response = self.client.get("/api/health")
        self.assertEqual(200, response.status_code)
        self.assertEqual("ok", response.json()["service"])
        self.assertEqual("ready", response.json()["knowledge"]["status"])

    def test_sqlserver_decimal_values_are_json_serializable(self) -> None:
        encoded = _json_safe(
            {
                "amount": Decimal("1234.56"),
                "whole_amount": Decimal("100"),
            }
        )
        self.assertEqual(1234.56, encoded["amount"])
        self.assertEqual(100, encoded["whole_amount"])

    def test_query_returns_route_plan_and_rows(self) -> None:
        """标准问题必须返回完整的路由证据、参数化计划与真实数据行。"""

        response = self.client.post(
            "/api/query",
            json={"question": "查询物料 110000012 的库存和库位", "limit": 20},
        )
        self.assertEqual(200, response.status_code)
        payload = response.json()
        self.assertEqual("completed", payload["status"])
        self.assertEqual("AiQueryPartOnHandV", payload["result"]["route"]["view_name"])
        self.assertEqual("物料编码", payload["result"]["column_labels"]["PartNum"])
        self.assertEqual("现有量", payload["result"]["column_labels"]["Qty"])
        self.assertTrue(any(row["Qty"] == 8290 for row in payload["result"]["rows"]))
        self.assertNotIn("dashboard", payload)

    def test_initial_dashboard_uses_current_profile(self) -> None:
        """初始 Dashboard 只读取当前画像，并提供真实概览图表。"""

        response = self.client.get("/api/dashboard")
        self.assertEqual(200, response.status_code)
        payload = response.json()
        self.assertEqual("initial", payload["mode"])
        self.assertGreaterEqual(len(payload["widgets"]), 4)
        self.assertEqual("hero", payload["widgets"][0]["size"])
        self.assertGreater(payload["widgets"][0]["value"], 0)
        self.assertTrue(any(widget["kind"] == "bar" for widget in payload["widgets"]))

    def test_fuzzy_query_requires_then_accepts_confirmation(self) -> None:
        """口语问题未确认时不查库，带回候选视图后才返回结果。"""

        question = "帮我看看物料 110000012 仓里还剩多少"
        pending = self.client.post("/api/query", json={"question": question})
        self.assertEqual("confirmation_required", pending.json()["status"])
        view_name = pending.json()["route"]["view_name"]
        confirmed = self.client.post(
            "/api/query",
            json={"question": question, "confirmed_view": view_name},
        )
        self.assertEqual("completed", confirmed.json()["status"])

    def test_dashboard_query_uses_separate_overview_chain(self) -> None:
        """Dashboard 对相对概念直接生成排序概况，不复用精确查询接口。"""

        fake = FakeLLMClient(
            {
                "status": "ready",
                "confidence": 0.78,
                "title": "库存较少物料",
                "summary": "按物料汇总各库位现有量并按升序展示较少项。",
                "route_reason": "用户想查看库存较少的物料概况",
                "source_views": ["AiQueryPartOnHandV"],
                "sql": (
                    "SELECT PartNum, PartDescription, SUM(Qty) AS TotalQty "
                    "FROM AiQueryPartOnHandV "
                    "GROUP BY PartNum, PartDescription "
                    "ORDER BY TotalQty ASC LIMIT 12"
                ),
                "parameters": [],
                "visualization": "ranking",
                "dimension_columns": ["PartDescription", "PartNum"],
                "metric_columns": ["TotalQty"],
                "display_units": {"TotalQty": "件"},
                "assumptions": ["将比较少理解为库存排序靠前的较少项"],
            }
        )
        with TestClient(create_test_app(llm_client=fake)) as client:
            response = client.post(
                "/api/dashboard/query",
                json={"question": "什么物料比较少"},
            )
        payload = response.json()
        self.assertEqual(200, response.status_code)
        self.assertEqual("completed", payload["status"])
        self.assertEqual("库存较少物料", payload["dashboard"]["title"])
        self.assertEqual("AiQueryPartOnHandV", payload["evidence"]["source_views"][0])
        self.assertEqual("bar", payload["dashboard"]["widgets"][0]["kind"])
        self.assertEqual("件", payload["dashboard"]["widgets"][0]["unit"])
        self.assertNotIn("result", payload)

    def test_catalog_exposes_curated_views_without_full_columns_by_default(self) -> None:
        """目录接口默认返回轻量行数画像，并保留 16 个已审核视图。"""

        response = self.client.get("/api/knowledge/catalog")
        self.assertEqual(200, response.status_code)
        payload = response.json()
        self.assertEqual(16, len(payload["views"]))
        self.assertEqual(16, len(payload["tables"]))
        self.assertEqual(12, len(payload["relationships"]))
        self.assertEqual(
            "docs/16张视图说明与字段关联分析.md",
            payload["semantic_source"]["document"],
        )
        self.assertTrue(payload["views"][0]["column_semantics"])
        self.assertNotIn("columns", payload["tables"][0])

    def test_knowledge_sync_runs_as_background_job(self) -> None:
        started = self.client.post(
            "/api/knowledge/sync",
            json={"reason": "api-test"},
        )
        self.assertEqual(200, started.status_code)
        job = started.json()
        self.assertIn(job["status"], {"queued", "running", "completed"})
        for _ in range(100):
            if job["status"] in {"completed", "failed"}:
                break
            time.sleep(0.01)
            response = self.client.get(f"/api/knowledge/sync/{job['job_id']}")
            self.assertEqual(200, response.status_code)
            job = response.json()
        self.assertEqual("completed", job["status"])


class AIApiTests(unittest.TestCase):
    """AI 接口验收：覆盖智能路由、自然语言回答与前端追问契约。"""

    def test_ai_routes_colloquial_inventory_and_answers_from_evidence(self) -> None:
        """AI 能理解口语库存问题，查询真实数据后再返回自然语言结论。"""

        fake = FakeLLMClient(
            {
                "status": "ready",
                "view_name": "AiQueryPartOnHandV",
                "confidence": 0.96,
                "intent_summary": "查询指定物料的库存数量",
                "route_reason": "用户询问物料当前存量",
                "clarification_question": None,
                "matched_concepts": ["物料", "存了多少"],
                "source_views": ["AiQueryPartOnHandV"],
                "sql": "SELECT PartNum, PartDescription, BinNum, BinName, Qty FROM AiQueryPartOnHandV WHERE PartNum = ? LIMIT 20",
                "parameters": ["110000012"],
                "assumptions": [],
            },
            answer="物料 110000012 当前查询到库存数量 8290。",
        )
        with TestClient(create_test_app(llm_client=fake)) as client:
            response = client.post(
                "/api/query",
                json={"question": "物料 110000012 存了多少"},
            )
        payload = response.json()
        self.assertEqual("completed", payload["status"])
        self.assertEqual("ai", payload["result"]["route"]["match_type"])
        self.assertTrue(payload["answer_generated_by_ai"])
        self.assertIn("8290", payload["answer"])
        self.assertTrue(any(row["Qty"] == 8290 for row in payload["result"]["rows"]))

    def test_ai_localizes_unknown_english_result_alias(self) -> None:
        """知识和固定词典均未覆盖的英文别名由当前模型补充中文表头。"""

        fake = ColumnLocalizingLLMClient(
            {
                "status": "ready",
                "confidence": 0.96,
                "intent_summary": "统计库存总数量",
                "route_reason": "库存视图包含现有量",
                "matched_concepts": ["库存", "合计"],
                "source_views": ["AiQueryPartOnHandV"],
                "sql": "SELECT SUM(Qty) AS total_qty FROM AiQueryPartOnHandV",
                "parameters": [],
                "assumptions": [],
            },
            answer="当前快照中的库存总数量为 **3728843.4834**。",
        )
        with TestClient(create_test_app(llm_client=fake)) as client:
            response = client.post(
                "/api/query",
                json={"question": "所有物料的库存总数量"},
            )

        payload = response.json()
        self.assertEqual("completed", payload["status"])
        self.assertEqual("库存总数量", payload["result"]["column_labels"]["total_qty"])
        self.assertEqual(1, fake.localization_calls)

    def test_ai_ambiguity_returns_frontend_clarification(self) -> None:
        """AI 无法唯一判断指标时不查库，而是返回可直接展示的追问。"""

        fake = FakeLLMClient(
            {
                "status": "clarification_required",
                "view_name": None,
                "confidence": 0.42,
                "intent_summary": "物料查询但指标不明确",
                "route_reason": "库存、采购和基础资料都可能符合",
                "clarification_question": "你想查物料的库存、采购价格，还是基础资料？",
                "clarification_kind": "choice",
                "clarification_options": ["查询库存", "查询采购价格", "查询基础资料"],
                "clarification_unit": None,
                "matched_concepts": ["物料"],
                "query_values": ["110000012"],
                "filter_constraints": [
                    {"column": "PartNum", "operator": "eq", "value": "110000012"}
                ],
                "requested_fields": [],
                "missing_information": ["需要查询的业务指标"],
                "operation": None,
                "metric_column": None,
            }
        )
        with TestClient(create_test_app(llm_client=fake)) as client:
            response = client.post(
                "/api/query",
                json={"question": "帮我查物料 110000012"},
            )
        payload = response.json()
        self.assertEqual("clarification_required", payload["status"])
        self.assertIn("库存", payload["analysis"]["clarification_question"])
        self.assertEqual("choice", payload["analysis"]["clarification_kind"])
        self.assertEqual(
            ["查询库存", "查询采购价格", "查询基础资料"],
            payload["analysis"]["clarification_options"],
        )

    def test_invalid_supplier_clarification_is_repaired_to_single_view_sql(self) -> None:
        """描述和值都已给出时，“想查询哪个供应商”是在索要答案，后端必须修复而非展示追问。"""

        fake = InvalidClarificationRepairingLLMClient(
            invalid_intent={
                "status": "clarification_required",
                "confidence": 0.9,
                "intent_summary": "齿条物料供应商查询",
                "route_reason": "误以为描述与供应商需要跨视图关联",
                "clarification_question": "您想查询哪个供应商？",
                "clarification_kind": "text",
                "matched_concepts": ["齿条", "供应商"],
                "source_views": ["AiQueryPoProgressV", "AiQueryPoOverViewV"],
                "filter_constraints": [
                    {
                        "column": "LineDesc",
                        "operator": "contains",
                        "value": "齿条,M5L1200H50S75",
                    }
                ],
                "requested_fields": ["VendorName"],
                "missing_information": ["供应商"],
            },
            repaired_intent={
                "status": "ready",
                "confidence": 0.98,
                "intent_summary": "查询齿条物料供应商",
                "route_reason": "采购进度视图同一行包含描述和供应商",
                "clarification_question": None,
                "matched_concepts": ["齿条", "供应商"],
                "source_views": ["AiQueryPoProgressV"],
                "filter_constraints": [
                    {
                        "column": "LineDesc",
                        "operator": "contains",
                        "value": "齿条,M5L1200H50S75",
                    }
                ],
                "requested_fields": ["VendorName"],
                "missing_information": [],
                "sql": "SELECT VendorName FROM AiQueryPoProgressV WHERE LineDesc LIKE ?",
                "parameters": ["%齿条,M5L1200H50S75%"],
                "assumptions": [],
            },
        )
        with TestClient(create_test_app(llm_client=fake)) as client:
            response = client.post(
                "/api/query",
                json={"question": "查询描述为齿条,M5L1200H50S75的供应商"},
            )

        payload = response.json()
        self.assertEqual("completed", payload["status"])
        self.assertEqual(2, fake.json_calls)
        self.assertEqual("AiQueryPoProgressV", payload["result"]["route"]["view_name"])
        self.assertIn("VendorName", payload["result"]["plan"]["base_sql"])
        self.assertIn("单个视图 AiQueryPoProgressV", fake.user_prompts[1])

    def test_hmi_description_query_does_not_require_code_clarification(self) -> None:
        """文本物料名可直接作为 BOM 描述包含条件，不应要求用户再提供物料编码。"""

        fake = InvalidClarificationRepairingLLMClient(
            invalid_intent={
                "status": "clarification_required",
                "confidence": 0.6,
                "intent_summary": "查询 HMI 上级部件",
                "route_reason": "误以为必须确认 HMI 是编码还是描述",
                "clarification_question": "HMI 是物料描述还是物料编码？",
                "clarification_kind": "choice",
                "clarification_options": ["物料描述", "物料编码"],
                "matched_concepts": ["HMI", "BOM上级部件"],
                "source_views": ["AiQueryBomV"],
                "filter_constraints": [
                    {
                        "column": "MtlPartDescription",
                        "operator": "contains",
                        "value": "HMI",
                    }
                ],
                "requested_fields": ["PartNum", "PartDescription"],
                "missing_information": ["物料编码"],
            },
            repaired_intent={
                "status": "ready",
                "confidence": 0.95,
                "intent_summary": "查询 HMI 上级部件",
                "route_reason": "BOM 视图支持按子物料描述查询上级部件",
                "clarification_question": None,
                "matched_concepts": ["HMI", "BOM上级部件"],
                "source_views": ["AiQueryBomV"],
                "filter_constraints": [
                    {
                        "column": "MtlPartDescription",
                        "operator": "contains",
                        "value": "HMI",
                    }
                ],
                "requested_fields": ["PartNum", "PartDescription"],
                "missing_information": [],
                "sql": "SELECT DISTINCT PartNum, PartDescription FROM AiQueryBomV WHERE MtlPartDescription LIKE ?",
                "parameters": ["%HMI%"],
                "assumptions": ["将 HMI 作为子物料描述关键词"],
            },
        )
        with TestClient(create_test_app(llm_client=fake)) as client:
            response = client.post(
                "/api/query",
                json={"question": "查询HMI这项物料属于哪个部件"},
            )

        payload = response.json()
        self.assertEqual("completed", payload["status"])
        self.assertEqual(2, fake.json_calls)
        self.assertEqual("AiQueryBomV", payload["result"]["route"]["view_name"])
        self.assertEqual(("%HMI%",), tuple(payload["result"]["plan"]["parameters"]))

    def test_clarification_history_prevents_reasking_answered_question(self) -> None:
        """多轮请求必须携带旧问答；模型重复提问时后端自动修复，不再打扰用户。"""

        fake = RepeatedClarificationLLMClient()
        with TestClient(create_test_app(llm_client=fake)) as client:
            response = client.post(
                "/api/query",
                json={
                    "question": "帮我查物料 110000012",
                    "clarification_history": [
                        {
                            "question": "你想查库存、采购价格，还是基础资料？",
                            "answer": "查询库存",
                        }
                    ],
                },
            )
        payload = response.json()
        self.assertEqual("completed", payload["status"])
        self.assertEqual(2, fake.json_calls)
        self.assertIn("查询库存", fake.user_prompts[0])
        self.assertIn("禁止重复询问", fake.user_prompts[0])
        self.assertNotIn("dashboard", payload)

    def test_ai_can_sum_inventory_without_material_identifier(self) -> None:
        """全局库存合计不应索要物料号，而应编译为受控 SUM 查询。"""

        fake = FakeLLMClient(
            {
                "status": "ready",
                "view_name": "AiQueryPartOnHandV",
                "confidence": 0.98,
                "intent_summary": "统计所有物料库存数量合计",
                "route_reason": "用户明确要求全部库存 Qty 合计",
                "clarification_question": None,
                "matched_concepts": ["所有物料", "库存数量合计"],
                "source_views": ["AiQueryPartOnHandV"],
                "sql": "SELECT SUM(Qty) AS Qty合计 FROM AiQueryPartOnHandV",
                "parameters": [],
                "assumptions": [],
                "display_units": {"Qty合计": "件"},
            },
            answer="当前快照中，所有物料的库存数量合计为 3728843.4834。",
        )
        with TestClient(create_test_app(llm_client=fake)) as client:
            response = client.post(
                "/api/query",
                json={"question": "所有物料的库存数量合计"},
            )
            payload = response.json()
            self.assertEqual("completed", payload["status"])
            self.assertEqual("text_to_sql", payload["result"]["plan"]["operation"])
            self.assertEqual(["AiQueryPartOnHandV"], payload["result"]["plan"]["source_views"])
            self.assertAlmostEqual(
                3728843.4834,
                payload["result"]["rows"][0]["Qty合计"],
            )
            self.assertFalse(payload["result"]["has_more"])
            download_id = payload["result"]["download_id"]
            self.assertTrue(download_id)

            download = client.get(f"/api/query/exports/{download_id}")
            self.assertEqual(200, download.status_code)
            workbook = load_workbook(BytesIO(download.content), read_only=True)
            worksheet = workbook["查询结果"]
            self.assertEqual(2, sum(1 for _ in worksheet.iter_rows()))
            workbook.close()

    def test_explicit_inventory_threshold_returns_preview_and_full_excel(self) -> None:
        """用户明确给出阈值 25 后，响应只含 500 行，但 Excel 必须包含全部结果。"""

        fake = FakeLLMClient(
            {
                "status": "ready",
                "confidence": 0.99,
                "intent_summary": "查询快没有库存的物料",
                "route_reason": "语义层规定按物料汇总后小于 25",
                "clarification_question": None,
                "matched_concepts": ["低库存"],
                "source_views": ["AiQueryPartOnHandV"],
                "sql": (
                    "SELECT PartNum, PartDescription, SUM(Qty) AS TotalQty "
                    "FROM AiQueryPartOnHandV GROUP BY PartNum, PartDescription "
                    "HAVING SUM(Qty) < 25 ORDER BY TotalQty ASC LIMIT 500"
                ),
                "parameters": [],
                "assumptions": [],
                "display_units": {"TotalQty": "件"},
                "limit_is_user_requested": False,
            },
            answer="库存量低于 **25** 的物料共有 **8712** 个。",
        )
        with TestClient(create_test_app(llm_client=fake)) as client:
            response = client.post(
                "/api/query",
                json={"question": "哪些物料库存数量低于 25", "limit": 500},
            )
            payload = response.json()
            self.assertEqual("completed", payload["status"])
            self.assertEqual(8712, payload["result"]["total_count"])
            self.assertEqual(500, len(payload["result"]["rows"]))
            self.assertTrue(payload["result"]["has_more"])
            self.assertNotIn("dashboard", payload)
            download_id = payload["result"]["download_id"]
            self.assertTrue(download_id)

            download = client.get(f"/api/query/exports/{download_id}")
            self.assertEqual(200, download.status_code)
            self.assertIn("spreadsheetml", download.headers["content-type"])
            workbook = load_workbook(BytesIO(download.content), read_only=True)
            worksheet = workbook["查询结果"]
            self.assertEqual(8713, sum(1 for _ in worksheet.iter_rows()))
            workbook.close()

    def test_ai_rewrites_answer_that_conflicts_with_nonempty_evidence(self) -> None:
        """SQL 有结果时若回答误称未找到，回答层必须自动重写一次。"""

        fake = FakeLLMClient(
            {
                "status": "ready",
                "confidence": 0.95,
                "intent_summary": "统计库存合计",
                "route_reason": "库存视图包含 Qty",
                "clarification_question": None,
                "matched_concepts": ["库存", "合计"],
                "source_views": ["AiQueryPartOnHandV"],
                "sql": "SELECT SUM(Qty) AS Qty合计 FROM AiQueryPartOnHandV",
                "parameters": [],
                "assumptions": [],
            },
            answer=[
                "库存合计为 3721955.4834，但未找到符合条件的数据。",
                "当前快照中的库存数量合计为 3721955.4834。",
            ],
        )
        with TestClient(create_test_app(llm_client=fake)) as client:
            response = client.post(
                "/api/query",
                json={"question": "所有物料库存合计"},
            )
        payload = response.json()
        self.assertEqual("completed", payload["status"])
        self.assertEqual(2, fake.text_calls)
        self.assertNotIn("未找到", payload["answer"])

    def test_invalid_intent_contract_becomes_natural_language_feedback(self) -> None:
        """模型结构连续异常时，接口仍返回回答状态，不向客户泄露技术命令。"""

        with TestClient(create_test_app(llm_client=BrokenIntentLLMClient())) as client:
            response = client.post(
                "/api/query",
                json={"question": "查询所有物料库存"},
            )
        payload = response.json()
        self.assertEqual(200, response.status_code)
        self.assertEqual("query_failed", payload["status"])
        self.assertTrue(payload["answer_generated_by_ai"])
        self.assertNotIn("JSON", payload["answer"])
        self.assertNotIn("DeepSeek", payload["answer"])
        self.assertNotIn("AiQuery", payload["answer"])

    def test_technical_failure_answer_is_replaced_by_safe_copy(self) -> None:
        """即使错误说明模型输出技术词，后端也必须替换为固定安全文案。"""

        with TestClient(create_test_app(llm_client=TechnicalFailureAnswerLLMClient())) as client:
            response = client.post(
                "/api/query",
                json={"question": "查询所有物料库存"},
            )
        payload = response.json()
        self.assertEqual("query_failed", payload["status"])
        self.assertFalse(payload["answer_generated_by_ai"])
        self.assertNotIn("DeepSeek", payload["answer"])
        self.assertNotIn("JSON", payload["answer"])
        self.assertNotIn("API", payload["answer"])

    def test_customer_mode_hides_technical_error_when_ai_is_not_configured(self) -> None:
        """客户模式缺少模型时返回自然语言状态，技术原因只留在后端日志。"""

        with TestClient(
            create_test_app(use_environment_ai=False, require_ai=True)
        ) as client:
            response = client.post(
                "/api/query",
                json={"question": "查询物料 110000012 的库存"},
            )
        payload = response.json()
        self.assertEqual(200, response.status_code)
        self.assertEqual("query_failed", payload["status"])
        self.assertNotIn("DeepSeek", payload["answer"])
        self.assertNotIn("未配置", payload["answer"])


if __name__ == "__main__":
    unittest.main()
